"""Extraction: xlsx program workbooks → structured rows; guidebook PDF → prose sections.

Checkpoint 1 findings this code encodes:
- The program tables live in the Excel workbooks; parsing is deterministic.
- Excel coerced rep/set ranges into dates ("8-10" → 2025-08-10); `unrange`
  reverses that exactly (month-day).
- The guidebook PDF is rasterized pages with a text overlay: prose extracts
  cleanly, but headings render letter-spaced ("T h e B e s t ..."). Image-only
  tables (screenshots) have no text layer and are skipped — the Excel covers them.
"""

from __future__ import annotations

import datetime
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pdfplumber

log = logging.getLogger(__name__)


# --------------------------------------------------------------------------- xlsx

@dataclass
class ProgramRow:
    program: str
    block: str
    week: int
    day: str
    exercise: str
    working_sets: int
    reps: str
    rpe_early: str
    rpe_last: str
    rest: str
    intensity: str | None
    warmup_sets: str | None
    substitutions: list[str] = field(default_factory=list)
    notes: str | None = None


def unrange(v: object) -> str | None:
    """Undo Excel's date coercion of ranges: datetime(2025, 8, 10) → "8-10"."""
    if isinstance(v, datetime.datetime):
        return f"{v.month}-{v.day}"
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _clean(v: object) -> str | None:
    if v is None:
        return None
    s = " ".join(str(v).split())
    return s or None


def parse_program_workbook(path: Path) -> tuple[list[ProgramRow], str]:
    """Parse one program workbook. Returns (rows, warmup_protocol_text).

    Layout (verified on both BTS workbooks): col B carries block titles
    ("Foundation Block"), week labels ("Week 1", same row as the header), and
    day labels (merged, so forward-filled); col C is the exercise; then
    intensity, warm-up sets, working sets, reps, 4 tracking columns, early RPE,
    last-set RPE, rest, two substitutions, notes.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    program = wb.sheetnames[0].replace("Program", "").strip()

    rows: list[ProgramRow] = []
    warmup_lines: list[str] = []
    block = week = day = None
    in_warmup = False

    for raw in ws.iter_rows(values_only=True):
        b, c = _clean(raw[1]), _clean(raw[2] if len(raw) > 2 else None)

        if b and "WARM UP" in b.upper():
            in_warmup = True
        if b and b.endswith("Block"):
            block, in_warmup = b.removesuffix("Block").strip(), False
            continue
        if in_warmup:
            cells = [_clean(v) for v in raw if _clean(v)]
            if cells:
                warmup_lines.append(" — ".join(cells))
            continue

        m = re.fullmatch(r"Week (\d+)", b or "")
        if m:
            week = int(m.group(1))
            continue  # header row
        if b:
            day = b
        if not c or c == "Exercise" or week is None:
            continue

        working = raw[5]
        if not isinstance(working, int):
            log.warning("%s: skipping row with non-int working sets: %r / %r", path.name, c, working)
            continue
        subs = [s for s in (_clean(raw[14]), _clean(raw[15])) if s]
        rows.append(ProgramRow(
            program=program, block=block or "", week=week, day=day or "",
            exercise=c, working_sets=working,
            reps=unrange(raw[6]) or "", rpe_early=unrange(raw[11]) or "",
            rpe_last=unrange(raw[12]) or "", rest=unrange(raw[13]) or "",
            intensity=None if (i := _clean(raw[3])) in (None, "N/A") else i,
            warmup_sets=unrange(raw[4]), substitutions=subs, notes=_clean(raw[16]),
        ))

    return rows, "\n".join(warmup_lines)


# ---------------------------------------------------------------------------- pdf

_SPACED = re.compile(r"^(?:\S ){3,}\S$")  # "T h e B e s t T r a i n i n g"


@dataclass
class GuideSection:
    title: str
    text: str
    first_page: int


def _normalize_heading(line: str) -> str:
    return "".join(line.split()) if _SPACED.match(line.strip()) else line.strip()


def parse_guidebook(path: Path) -> list[GuideSection]:
    """Split the guidebook into sections at heading boundaries.

    Headings are either full title pages (few words, all caps) or letter-spaced
    lines within a page. Per-page failures are logged and skipped, never fatal.
    """
    sections: list[GuideSection] = []
    title, buf, first = "Introduction", [], 1

    def flush(next_title: str, page: int) -> None:
        nonlocal title, buf, first
        text = "\n".join(buf).strip()
        if len(text.split()) > 30:
            sections.append(GuideSection(title=title, text=text, first_page=first))
        title, buf, first = next_title, [], page

    with pdfplumber.open(path) as pdf:
        for pageno, page in enumerate(pdf.pages, 1):
            try:
                text = page.extract_text(x_tolerance=1.5) or ""
            except Exception:
                log.warning("%s: page %d failed text extraction; skipped", path.name, pageno)
                continue
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            # Drop running footer
            lines = [ln for ln in lines
                     if not ln.upper().startswith("THE BODYBUILDING TRANSFORMATION SYSTEM")]
            if not lines:
                continue
            body_words = sum(len(ln.split()) for ln in lines)
            joined = " ".join(lines)
            if body_words <= 8 and joined.upper() == joined:  # title page
                flush(_normalize_heading(joined).title() if _SPACED.match(joined) else joined.title(), pageno)
                continue
            for ln in lines:
                if _SPACED.match(ln):
                    flush(_camel_title(_normalize_heading(ln)), pageno)
                else:
                    buf.append(ln)
        flush("", len(pdf.pages))
    return sections


def _camel_title(collapsed: str) -> str:
    """"TheBestTrainingSplit" → "The Best Training Split"; all-caps stays as-is."""
    if collapsed.upper() == collapsed:
        return collapsed.title()
    return re.sub(r"(?<=[a-z])(?=[A-Z])", " ", collapsed)
